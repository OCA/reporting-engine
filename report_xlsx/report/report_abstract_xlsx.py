# Copyright 2015 ACSONE SA/NV (<http://acsone.eu>)
# License AGPL-3.0 or later (https://www.gnu.org/licenses/agpl.html).

import logging
import re
from io import BytesIO

import openpyxl

from odoo import _, api, models
from odoo.exceptions import ValidationError
from odoo.modules.module import get_module_resource
from odoo.tools import safe_eval

from .utils.xlsx_utils import HORIZ_ALIG_ACPTD_VALS, VERT_ALIG_ACPTD_VALS

_logger = logging.getLogger(__name__)

try:
    import xlsxwriter

    class PatchedXlsxWorkbook(xlsxwriter.Workbook):
        def _check_sheetname(self, sheetname, is_chartsheet=False):
            """We want to avoid duplicated sheet names exceptions the same following
            the same philosophy that Odoo implements overriding the main library
            to avoid the 31 characters limit triming the strings before sending them
            to the library.

            In some cases, there's not much control over this as the reports send
            automated data and the potential exception is hidden underneath making it
            hard to debug the original issue. Even so, different names can become the
            same one as their strings are trimmed to those 31 character limit.

            This way, once we come across with a duplicated, we set that final 3
            characters with a sequence that we evaluate on the fly. So for instance:

            - 'Sheet name' will be 'Sheet name~01'
            - The next 'Sheet name' will try to rename to 'Sheet name~01' as well and
              then that will give us 'Sheet name~02'.
            - And the next 'Sheet name' will try to rename to 'Sheet name~01' and then
              to 'Sheet name~02' and finally it will be able to 'Sheet name~03'.
            - An so on as many times as duplicated sheet names come to the workbook up
              to 100 for each sheet name. We set such limit as we don't want to truncate
              the strings too much and keeping in mind that this issue don't usually
              ocurrs.
            """
            try:
                return super()._check_sheetname(sheetname, is_chartsheet=is_chartsheet)
            except xlsxwriter.exceptions.DuplicateWorksheetName:
                pattern = re.compile(r"~[0-9]{2}$")
                duplicated_secuence = (
                    re.search(pattern, sheetname) and int(sheetname[-2:]) or 0
                )
                # Only up to 100 duplicates
                deduplicated_secuence = "~{:02d}".format(duplicated_secuence + 1)
                if duplicated_secuence > 99:
                    raise xlsxwriter.exceptions.DuplicateWorksheetName
                if duplicated_secuence:
                    sheetname = re.sub(pattern, deduplicated_secuence, sheetname)
                elif len(sheetname) <= 28:
                    sheetname += deduplicated_secuence
                else:
                    sheetname = sheetname[:28] + deduplicated_secuence
            # Refeed the method until we get an unduplicated name
            return self._check_sheetname(sheetname, is_chartsheet=is_chartsheet)

    # "Short string"

    xlsxwriter.Workbook = PatchedXlsxWorkbook

except ImportError:
    _logger.debug("Can not import xlsxwriter`.")


class ReportXlsxAbstract(models.AbstractModel):
    _name = "report.report_xlsx.abstract"
    _description = "Abstract XLSX Report"

    # Relative path based on module location
    _boilerplate_template_file_path = None

    def _get_objs_for_report(self, docids, data):
        """
        Returns objects for xlsx report.  From WebUI these
        are either as docids taken from context.active_ids or
        in the case of wizard are in data.  Manual calls may rely
        on regular context, setting docids, or setting data.

        :param docids: list of integers, typically provided by
            qwebactionmanager for regular Models.
        :param data: dictionary of data, if present typically provided
            by qwebactionmanager for TransientModels.
        :param ids: list of integers, provided by overrides.
        :return: recordset of active model for ids.
        """
        if docids:
            ids = docids
        elif data and "context" in data:
            ids = data["context"].get("active_ids", [])
        else:
            ids = self.env.context.get("active_ids", [])
        return self.env[self.env.context.get("active_model")].browse(ids)

    def _report_xlsx_currency_format(self, currency):
        """Get the format to be used in cells (symbol included).
        Used in account_financial_report addon"""
        s_before = currency.symbol if currency.position == "before" else ""
        s_after = " %s" % currency.symbol if currency.position == "after" else ""
        return f"{f'{s_before}'}#,##0.{'0' * currency.decimal_places}{f'{s_after}'}"

    def create_xlsx_report(self, docids, data):
        objs = self._get_objs_for_report(docids, data)
        file_data = BytesIO()
        if self._boilerplate_template_file_path is not None and data.get("data", False):
            workbook = self._get_boilerplate_template(file_data, data)
        else:
            workbook = self._get_new_workbook(file_data)
        self.generate_xlsx_report(workbook, data, objs)
        workbook.close()
        file_data.seek(0)
        return file_data.read(), "xlsx"

    def get_workbook_options(self):
        """
        See https://xlsxwriter.readthedocs.io/workbook.html constructor options
        :return: A dictionary of options
        """
        return {}

    def generate_xlsx_report(self, workbook, data, objs):
        raise NotImplementedError()

    def _get_new_workbook(self, file_data):
        """
        :return: empty Workbook
        :rtype: xlsxwriter.Workbook object
        """
        return xlsxwriter.Workbook(file_data, self.get_workbook_options())

    # flake8: noqa: C901
    @api.model
    def _copy_cell_format(self, xslx, cell):
        """
        :return: a format object that needs to be applied coming from the
        openyxl.cell.cell.Cell object
        :rtype: xlsxwriter.format.Format
        """
        cell_format = xslx.add_format()
        values_dict = dict()
        bs_field = "border_style"
        if hasattr(cell.fill.bgColor, "rgb") and cell.fill.bgColor.value != "00000000":
            bg_color = "#" + cell.fill.bgColor.rgb[2:]
            values_dict.setdefault("bg_color", bg_color)
        if cell.font:
            if hasattr(cell.font.color, "rgb"):
                font_color = "#" + cell.font.color.rgb[2:]
                values_dict.setdefault("font_color", font_color)
            if cell.font.b:
                values_dict.setdefault("bold", cell.font.b)
            if cell.font.i:
                values_dict.setdefault("italic", cell.font.i)
            if cell.font.u:
                values_dict.setdefault("underline", cell.font.u)
        if cell.alignment:
            if cell.alignment.wrapText:
                values_dict.setdefault("text_wrap", cell.alignment.wrapText)
            values_dict.setdefault("align", list())
            if cell.alignment.horizontal in HORIZ_ALIG_ACPTD_VALS.keys():
                values_dict["align"].append(
                    HORIZ_ALIG_ACPTD_VALS.get(cell.alignment.horizontal)
                )
            if cell.alignment.vertical in VERT_ALIG_ACPTD_VALS.keys():
                values_dict["align"].append(
                    VERT_ALIG_ACPTD_VALS.get(cell.alignment.vertical)
                )
        for side in ["left", "right", "top", "bottom"]:
            if hasattr(cell.border, side) and hasattr(
                getattr(cell.border, side), bs_field
            ):
                values_dict.setdefault(side, 1)
        for key, value in values_dict.items():
            func_name = "set_%s" % key
            if hasattr(cell_format, func_name):
                if isinstance(value, list):
                    if not value:
                        continue
                    values_to_assign = value
                else:
                    values_to_assign = [value]
                for val in values_to_assign:
                    getattr(cell_format, func_name)(val)
        return cell_format

    def _copy_xlsx(self, file_data, template_xlsx):
        """
        :return: a copy of the openpyxl.Workbook object on a xlsxwriter.Workbook object. Converts all the content from one type of object to the other
        :rtype: xlsxwriter.Workbook object
        """
        new_xlsx = xlsxwriter.Workbook(file_data, self.get_workbook_options())
        template_sheets = template_xlsx.get_sheet_names()
        for sheet_name in template_sheets:
            new_xlsx.add_worksheet(sheet_name)
        for sheet in template_sheets:
            openpyxl_active_sheet = template_xlsx.get_sheet_by_name(sheet)
            xlsxwriter_active_sheet = new_xlsx.get_worksheet_by_name(sheet)
            for count, row in enumerate(openpyxl_active_sheet.rows):
                for cell in row:
                    if isinstance(cell, openpyxl.cell.cell.MergedCell):
                        continue
                    else:
                        # 1. Set Column Width
                        column_index = cell.column - 1
                        cell_width = openpyxl_active_sheet.column_dimensions[
                            cell.column_letter
                        ].width
                        xlsxwriter_active_sheet.set_column(
                            column_index, column_index, cell_width
                        )
                        # 2. Set Cell Format for each cell
                        cell_format = self._copy_cell_format(new_xlsx, cell)
                        xlsxwriter_active_sheet.write(
                            cell.coordinate, cell.value, cell_format
                        )
                # 3. Set Row Height for each row
                row_index = cell.row - 1
                cell_height = openpyxl_active_sheet.row_dimensions[cell.row].height
                xlsxwriter_active_sheet.set_row(row_index, cell_height)
            # 4. Merge merged cells at the end
            for merge_range in openpyxl_active_sheet.merged_cells:
                xlsxwriter_active_sheet.merge_range(merge_range.coord, "")
        return new_xlsx

    def _get_boilerplate_template(self, file_data, data):
        """
        :return: copy of the Boilerplate Template of the report if everything is
        correctly set up, blank workbook otherwise
        :rtype: xlsxwriter.Workbook object
        """
        report_whole_path = safe_eval(data.get("data"))[0]
        report_path = [elem for elem in report_whole_path.split("/") if "." in elem]
        report_path = report_path[0] if report_path else False
        module, dummy = tuple(report_path.split("."))
        module_path, file_name = self._boilerplate_template_file_path.rsplit("/", 1)
        file_path = get_module_resource(module, module_path, file_name)
        if not file_path:
            raise ValidationError(
                _(
                    "Boilerplate Template file path not properly defined: %s."
                    % self._boilerplate_template_file_path
                )
            )
        try:
            boilerplate_template = openpyxl.load_workbook(file_path)
            return self._copy_xlsx(file_data, boilerplate_template)
        except Exception as e:
            _logger.exception(e)
            return self._get_new_workbook(file_data)
